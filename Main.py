from tkinter import *
from tkinter import PhotoImage
from tkinter.messagebox import showinfo, showerror, showwarning
from openpyxl import load_workbook, Workbook
import os

class MainMenu:
    def __init__(self):
        self.window = Tk()
        self.window.geometry("1024x768")
        self.window.title("Glow Getter")
        
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        window_width = 1024
        window_height = 768
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        self.window.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
        self.window.resizable(False, False)
        
        self.background = PhotoImage(file="assets/background.png")
        self.button_image = PhotoImage(file="assets/login.png")
        self.button1_image = PhotoImage(file="assets/register.png")
        
        self.background_label = Label(self.window, image=self.background)
        self.background_label.pack()
        
        self.username_entry = Entry(self.window, width=30)
        self.username_entry.place(x=389, y=352)
        self.username_entry.config(font=("Arial"))
        
        self.password_entry = Entry(self.window, width=30, show="*")
        self.password_entry.place(x=389, y=437)
        self.password_entry.config(font=("Arial"))
        
        self.login_button = Button(self.window, image=self.button_image, command=self.login, border=0)
        self.login_button.place(x=440, y=512)
        
        
        self.window.mainloop()
    
    
    def login(self):
        
        username = self.username_entry.get()
        password = self.password_entry.get()
        if username == '' or password == '' and username == ' ' or password == ' ':
            showerror("Error", "Please enter a valid username and password")
        else:
            self.save_login(username, password)
            self.window.destroy()
            ShoppingWindow(username)
    def save_login(self, username, password):
        file_path = 'database/customers.xlsx'
        
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Username', 'Password'])
            
        sheet.append([username, password])
        wb.save(file_path)
        

class ShoppingWindow():
    def __init__(self, username):
        self.window = Tk()
        self.window.geometry("1280x960")
        self.window.title("Glow Getter")
        
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        window_width = 1280
        window_height = 960
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        self.window.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
        self.window.resizable(False, False)
        
        self.shop_background = PhotoImage(file="assets/shop_background.png")
        self.background_label1 = Label(self.window, image=self.shop_background)
        self.background_label1.pack()
        
        message_butt = PhotoImage(file="assets/message.png")
        self.message_button = Button(self.window, image=message_butt, border=0)
        self.message_button.config(bg='#939974', border=0, activebackground='#939974', cursor="hand2", command=self.send_message)
        self.message_button.place(x=1150, y=42)
        
        self.greeting = Label(self.window, text=f"Welcome, {username}")
        self.greeting.place(x=140, y=30, width=1001, height=100)
        self.greeting.config(font=("Arial",35, "bold"), bg='#939974', fg='white')
        
        self.load_products("products.xlsx")
             
        self.window.mainloop()
    
    def send_message(self):
        ChatWindow()
        
    def load_products(self, filename):
        pass

class ChatWindow:
    def __init__(self):
        self.window = Tk()
        self.window.geometry("600x400")
        self.window.title("Glow Getter : Message Staff")
        self.window.resizable(False, False)
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        window_width = 600
        window_height = 400
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        self.window.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
        self.window.resizable(False, False)
        
        self.chat_background1 = PhotoImage(file="assets/chat_background.png")
        self.background_label2 = Label(self.window, image=self.chat_background1)
        self.background_label2.pack()
        
        self.chat_box = Text(self.window, width=70, height=15, state=DISABLED)
        self.chat_box.pack(pady=10)
        
        self.chat_entry = Entry(self.window, width=70)
        self.chat_entry.pack(side=LEFT, padx=10)
        
        self.send_button = Button(self.window, text="Send")
        self.send_button.pack(pady=10)
        
        self.window.mainloop()


if __name__ == "__main__":
    MainMenu()  
    